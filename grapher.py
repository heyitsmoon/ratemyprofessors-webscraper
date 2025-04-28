import pandas as pd
import openpyxl
#import pt_creator
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList


def QD_overtime(df, sheet, dashboard):

    df["Review Date"] = pd.to_datetime(df["Review Date"])

    df["Year"] = df["Review Date"].dt.year

    # Calculate average quality and difficulty by year
    avg_quality_by_year = df.groupby('Year')['Quality'].mean().reset_index()
    avg_difficulty_by_year = df.groupby('Year')['Difficulty'].mean().reset_index()

    # Merge the quality and difficulty data into a single DataFrame
    avg_data = pd.merge(avg_quality_by_year, avg_difficulty_by_year, on='Year', suffixes=('_quality', '_difficulty'))
    line_chart = LineChart()
    line_chart.title = "Quality/Difficulty Over Time"
    line_chart.x_axis.title = "Year"
    line_chart.y_axis.title = "Average"
    line_chart.width = 33.9
    line_chart.y_axis.majorUnit = 0.5


    # Define data range for average quality and difficulty
    x_values = Reference(sheet, min_col=5, min_row=3, max_row=len(avg_data)+3)
    y_values1 = Reference(sheet, min_col=6, min_row=2, max_row=len(avg_data)+2)  
    y_values2 = Reference(sheet, min_col=7, min_row=2, max_row=len(avg_data)+2) 

    # Add data to the chart
    line_chart.add_data(y_values1, titles_from_data=True)
    line_chart.add_data(y_values2, titles_from_data=True)
    line_chart.set_categories(x_values)

    line_chart.legend = None

    # Append the chart to the worksheet
    dashboard.add_chart(line_chart, "A15")



def Rating(sheet,dashboard):

    bar_chart = BarChart()
    #bar_chart.title = "Frequency of Difficulty Ratings"
    bar_chart.y_axis.title = "Count"
    bar_chart.x_axis.title = "Difficulty Rating"

    # Define data for the bar chart
    cats = Reference(sheet, min_col=1, min_row=2, max_row=6)
    data = Reference(sheet, min_col=2, min_row=2, max_row=6)
    

    bar_chart.add_data(data)
    bar_chart.set_categories(cats)
    
    # Remove legend
    bar_chart.legend = None
    bar_chart.width  = 8.5
    bar_chart.height = 7.4  # Set the height to 10 Excel rows
    # Add the bar chart to the Excel sheet
    dashboard.add_chart(bar_chart, "A1")


def Textbook(sheet, dashboard):

    
    bar_chart = BarChart()
    bar_chart.y_axis.title = "Count"
    bar_chart.x_axis.title = "Textbook Required"

    cats = Reference(sheet, min_col=5,min_row=2,max_row=3)
    data = Reference(sheet, min_col=6, min_row=2, max_row=3)
    

    bar_chart.add_data(data)
    bar_chart.set_categories(cats)

    bar_chart.legend = None
    bar_chart.height = 7.4
    bar_chart.width  = 8.5
    bar_chart.y_axis.scaling.min = 0
    
    
    dashboard.add_chart(bar_chart,"K1")


def GradeReported(sheet,dashboard):

    bar_chart = BarChart()
    bar_chart.y_axis.title = "Count"
    bar_chart.x_axis.title = "Grades"
    
    cats = Reference(sheet, min_col=7, min_row=2, max_row=5) #used to be 14
    data = Reference(sheet, min_col=8, min_row=2, max_row=5) #used to be 14

    bar_chart.add_data(data)
    bar_chart.set_categories(cats)

    bar_chart.dataLabels = DataLabelList()
    bar_chart.dataLabels.showVal = True
    bar_chart.legend = None
    bar_chart.height = 7.4
    bar_chart.width  = 8.5

    dashboard.add_chart(bar_chart, "P1")

def Attendance(sheet, dashboard):
    
    bar_chart = BarChart()
    bar_chart.y_axis.title = "Count"
    bar_chart.x_axis.title = "Attendance"

    cats = Reference(sheet, min_col=3, min_row=2, max_row=3)
    data = Reference(sheet, min_col=4, min_row=2, max_row=3)
    

    bar_chart.add_data(data)
    bar_chart.set_categories(cats)

    bar_chart.legend = None
    bar_chart.height = 7.4
    bar_chart.width  = 8.5
    
    dashboard.add_chart(bar_chart, "F1")

    
def graphit(professor_name):
    # Read data from CSV using pandas
    df = pd.read_csv(rf'{professor_name} Reviews.csv')

    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    data = wb.active
    data.title = 'Data'
    title_row = ['professor_name','TakeAgain%','Overall_Difficulty','Rating','Date','Course', 'Review Date', 'Review Year', 'Quality', 'Difficulty', 'For Credit', 'Attendance', 'Would Take Again', 'Grade', 'Textbook', 'Online Class', 'Comment']
    data.append(title_row)
    
    for i, row in enumerate(df.iterrows(), start=2):
        for j, value in enumerate(row[1], start=1):
            if ((j==11 or j==13 or j==14) and value=="Helpful"):
                continue
            data.cell(row=i, column=j, value=value)

    sheet = wb.create_sheet(title='Pivot Tables')
    dashboard = wb.create_sheet(title='Dashboard')
    
    QD_overtime(df, sheet, dashboard)
    Rating(sheet, dashboard)
    Textbook(sheet, dashboard)
    GradeReported(sheet, dashboard)
    Attendance(sheet, dashboard)

    
    wb.save(f"{professor_name} Reviews.xlsx")
    
    #pt_creator.create_pivots(f"{professor_name} Reviews.xlsx")
